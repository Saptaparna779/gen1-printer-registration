"""
Converts a pytest JUnit XML report + the corresponding .feature file into
a human-readable Excel report: one row per scenario, with pass/fail,
duration, and the plain-language scenario name.

Usage:
    python generate_cucumber_excel_report.py GOAR-3
"""
import sys
import xml.etree.ElementTree as ET
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def parse_feature_scenarios(feature_path):
    """Extract scenario names in order from a .feature file."""
    with open(feature_path, "r", encoding="utf-8") as f:
        content = f.read()
    return re.findall(r"^\s*Scenario:\s*(.+)$", content, re.MULTILINE)


def parse_junit_results(xml_path):
    """Extract test results in order from a JUnit XML report."""
    tree = ET.parse(xml_path)
    results = []
    for testcase in tree.iter("testcase"):
        name = testcase.get("name")
        time = float(testcase.get("time", 0))
        failure = testcase.find("failure")
        error = testcase.find("error")
        skipped = testcase.find("skipped")
        if failure is not None:
            status = "FAILED"
        elif error is not None:
            status = "ERROR"
        elif skipped is not None:
            status = "SKIPPED"
        else:
            status = "PASSED"
        results.append({"name": name, "status": status, "time": time})
    return results


def main():
    if len(sys.argv) != 2:
        print("Usage: python generate_cucumber_excel_report.py <ISSUE-KEY>")
        sys.exit(1)

    issue_key = sys.argv[1].upper()
    feature_path = f"tests/features/{issue_key}.feature"
    xml_path = f"reports/{issue_key}_cucumber_results.xml"
    output_path = f"reports/{issue_key}_cucumber_execution_report.xlsx"

    scenarios = parse_feature_scenarios(feature_path)
    results = parse_junit_results(xml_path)

    if len(scenarios) != len(results):
        print(
            f"WARNING: {len(scenarios)} scenarios found in {feature_path} but "
            f"{len(results)} test results found in {xml_path}. Rows will still "
            "be written in order, but double-check they align correctly."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{issue_key} Cucumber Results"

    headers = ["#", "Scenario", "Status", "Duration (s)"]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for i, result in enumerate(results):
        scenario_name = scenarios[i] if i < len(scenarios) else "(scenario name unavailable)"
        row = [i + 1, scenario_name, result["status"], round(result["time"], 3)]
        ws.append(row)
        status_cell = ws.cell(row=i + 2, column=3)
        status_cell.fill = pass_fill if result["status"] == "PASSED" else fail_fill
        status_cell.font = Font(bold=True)
        status_cell.alignment = Alignment(horizontal="center")

    passed_count = sum(1 for r in results if r["status"] == "PASSED")
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"{passed_count}/{len(results)} scenarios passed")

    ws.column_dimensions[get_column_letter(1)].width = 5
    ws.column_dimensions[get_column_letter(2)].width = 90
    ws.column_dimensions[get_column_letter(3)].width = 12
    ws.column_dimensions[get_column_letter(4)].width = 14
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Report saved to {output_path}")


if __name__ == "__main__":
    main()
